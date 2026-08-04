"""
src/importer/fetchers/amfi_flows_fetcher.py
───────────────────────────────────────────
Fetches AMFI Category-Wise Monthly Flows (Dataset B) + AUM (Dataset C)
from the amfiindia.com website.

Three-layer probe/fallback strategy
------------------------------------
Layer 1 — Strapi CMS API probe
    AMFI migrated to Next.js + Strapi in 2024. The old /spages/*.xlsx URLs
    return 404. This layer probes the JS bundle for the CMS hostname and
    calls the Strapi REST API directly.

Layer 2 — HTML scrape (httpx + BeautifulSoup)
    Parses <table> elements from the rendered page for each target month.

Layer 3 — Manual URL override
    Reads AMFI_EXCEL_URL from the environment. Allows a one-time manual
    override once a known-good direct download URL is identified.

If all layers return empty, the fetcher logs a WARNING and returns [] without
crashing the import pipeline.

Watermark key:
    source   = "amfi_category_flows"
    symbol   = "INDUSTRY"
    dataset  = "flows"

Each returned dict has keys:
    report_month        (date)   First day of the AMFI report month
    category_name       (str)    Verbatim AMFI category label
    subcategory_group   (str)    Normalised: Equity | Debt | Hybrid | Passive | FoF | Solution | Other
    gross_purchase_cr   (float)  Gross purchases ₹ Crore
    gross_redemption_cr (float)  Gross redemptions ₹ Crore
    net_flow_cr         (float)  gross_purchase − gross_redemption (computed in Python)
    closing_aum_cr      (float)  Month-end AUM ₹ Crore
    flow_pct_of_aum     (float)  net_flow_cr / closing_aum_cr × 100 (computed in Python)
"""

from __future__ import annotations

import logging
import os
import re
import time
from datetime import date, datetime
from typing import Optional

logger = logging.getLogger(__name__)

# ── Subcategory normalisation map ─────────────────────────────────────────────
# Verbatim AMFI category name → normalised subcategory_group
# Category names change slightly over time; we use fuzzy prefix matching as
# a fallback (see _normalize_subcategory).

_SUBCATEGORY_MAP: dict[str, str] = {
    # ── Equity ────────────────────────────────────────────────────────────────
    "Large Cap Fund":                                "Equity",
    "Large & Mid Cap Fund":                          "Equity",
    "Large and Mid Cap Fund":                        "Equity",
    "Mid Cap Fund":                                  "Equity",
    "Small Cap Fund":                                "Equity",
    "Multi Cap Fund":                                "Equity",
    "Flexi Cap Fund":                                "Equity",
    "Focused Fund":                                  "Equity",
    "Value Fund":                                    "Equity",
    "Contra Fund":                                   "Equity",
    "ELSS":                                          "Equity",
    "Sectoral/Thematic Funds":                       "Equity",
    "Sectoral/ Thematic Funds":                      "Equity",
    "Dividend Yield Fund":                           "Equity",
    # ── Debt ──────────────────────────────────────────────────────────────────
    "Overnight Fund":                                "Debt",
    "Liquid Fund":                                   "Debt",
    "Ultra Short Duration Fund":                     "Debt",
    "Low Duration Fund":                             "Debt",
    "Money Market Fund":                             "Debt",
    "Short Duration Fund":                           "Debt",
    "Medium Duration Fund":                          "Debt",
    "Medium to Long Duration Fund":                  "Debt",
    "Long Duration Fund":                            "Debt",
    "Dynamic Bond":                                  "Debt",
    "Corporate Bond Fund":                           "Debt",
    "Credit Risk Fund":                              "Debt",
    "Banking and PSU Fund":                          "Debt",
    "Banking and PSU Debt Fund":                     "Debt",
    "Gilt Fund":                                     "Debt",
    "Gilt Fund with 10 year constant duration":      "Debt",
    "Gilt Fund with 10 Year Constant Duration":      "Debt",
    "Floater Fund":                                  "Debt",
    # ── Hybrid ────────────────────────────────────────────────────────────────
    "Conservative Hybrid Fund":                      "Hybrid",
    "Balanced Hybrid Fund":                          "Hybrid",
    "Aggressive Hybrid Fund":                        "Hybrid",
    "Dynamic Asset Allocation or Balanced Advantage":"Hybrid",
    "Dynamic Asset Allocation":                      "Hybrid",
    "Balanced Advantage":                            "Hybrid",
    "Multi Asset Allocation":                        "Hybrid",
    "Arbitrage Fund":                                "Hybrid",
    "Equity Savings":                                "Hybrid",
    "Equity Savings Fund":                           "Hybrid",
    # ── Passive ───────────────────────────────────────────────────────────────
    "Index Funds":                                   "Passive",
    "ETFs":                                          "Passive",
    "Gold ETFs":                                     "Passive",
    "Other ETFs":                                    "Passive",
    "International ETFs":                            "Passive",
    # ── Fund of Funds ─────────────────────────────────────────────────────────
    "Fund of Funds (Domestic)":                      "FoF",
    "Fund of Funds (Overseas)":                      "FoF",
    "Fund of Funds- Domestic":                       "FoF",
    "Fund of Funds- Overseas":                       "FoF",
    # ── Solution-oriented ────────────────────────────────────────────────────
    "Retirement Fund":                               "Solution",
    "Children's Fund":                               "Solution",
}

# Fuzzy prefix tokens for normalisation fallback (order matters — more specific first)
_FUZZY_MAP: list[tuple[str, str]] = [
    ("gold etf", "Passive"),
    ("index fund", "Passive"),
    ("etf", "Passive"),
    ("liquid", "Debt"),
    ("overnight", "Debt"),
    ("gilt", "Debt"),
    ("ultra short", "Debt"),
    ("low duration", "Debt"),
    ("money market", "Debt"),
    ("short duration", "Debt"),
    ("medium", "Debt"),
    ("long duration", "Debt"),
    ("dynamic bond", "Debt"),
    ("corporate bond", "Debt"),
    ("credit risk", "Debt"),
    ("banking and psu", "Debt"),
    ("floater", "Debt"),
    ("large cap", "Equity"),
    ("large & mid cap", "Equity"),
    ("mid cap", "Equity"),
    ("small cap", "Equity"),
    ("multi cap", "Equity"),
    ("flexi cap", "Equity"),
    ("focused", "Equity"),
    ("value fund", "Equity"),
    ("contra", "Equity"),
    ("elss", "Equity"),
    ("sectoral", "Equity"),
    ("thematic", "Equity"),
    ("dividend yield", "Equity"),
    ("conservative hybrid", "Hybrid"),
    ("balanced hybrid", "Hybrid"),
    ("aggressive hybrid", "Hybrid"),
    ("balanced advantage", "Hybrid"),
    ("dynamic asset allocation", "Hybrid"),
    ("multi asset", "Hybrid"),
    ("arbitrage", "Hybrid"),
    ("equity savings", "Hybrid"),
    ("fund of funds", "FoF"),
    ("retirement", "Solution"),
    ("children", "Solution"),
]


def _normalize_subcategory(category_name: str) -> str:
    """Exact match first; then fuzzy prefix match. Falls back to 'Other'."""
    if category_name in _SUBCATEGORY_MAP:
        return _SUBCATEGORY_MAP[category_name]
    lc = category_name.lower()
    for token, group in _FUZZY_MAP:
        if token in lc:
            return group
    return "Other"


def _parse_amount(value: str) -> float:
    """
    Parse AMFI flow/AUM string like '12,345.67' or '(3,210.44)' → float.
    Negative values may be represented as (N) in older Excel formats.
    """
    if not value or str(value).strip() in ("", "-", "N.A.", "NA"):
        return 0.0
    s = str(value).strip().replace(",", "").replace(" ", "")
    negative = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        result = float(s)
        return -result if negative else result
    except ValueError:
        return 0.0


def _month_dates(months_back: int) -> list[date]:
    """
    Return a list of first-of-month dates going back `months_back` months from today.
    Most recent first.
    """
    today = date.today()
    dates = []
    year, month = today.year, today.month
    for _ in range(months_back):
        dates.append(date(year, month, 1))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return dates


# ── Layer 1: AMFI Discovered Direct Portal SIF Excel Links ────────────────────

_PORTAL_VOL_MAP: list[tuple[int, list[tuple[str, date]]]] = [
    (25, [("issueI", date(2025,6,1)), ("issueII", date(2025,9,1)), ("issueIII", date(2025,12,1)), ("issueIV", date(2026,3,1))]),
    (24, [("issueI", date(2024,6,1)), ("issueII", date(2024,9,1)), ("issueIII", date(2024,12,1)), ("issueIV", date(2025,3,1))]),
    (23, [("issueI", date(2023,6,1)), ("issueII", date(2023,9,1)), ("issueIII", date(2023,12,1)), ("issueIV", date(2024,3,1))]),
    (22, [("issueI", date(2022,6,1)), ("issueII", date(2022,9,1)), ("issueIII", date(2022,12,1)), ("issueIV", date(2023,3,1))]),
    (21, [("issueI", date(2021,6,1)), ("issueII", date(2021,9,1)), ("issueIII", date(2021,12,1)), ("issueIV", date(2022,3,1))]),
    (20, [("issueI", date(2020,6,1)), ("issueII", date(2020,9,1)), ("issueIII", date(2020,12,1)), ("issueIV", date(2021,3,1))]),
    (19, [("issueI", date(2019,6,1)), ("issueII", date(2019,9,1)), ("issueIII", date(2019,12,1)), ("issueIV", date(2020,3,1))]),
]


def _fetch_via_portal_statistical_reports(target_months: list[date]) -> list[dict]:
    """
    Fetch and parse direct monthly SIF statistical Excel files
    from portal.amfiindia.com/spages/am<mon><year>repo.xls
    """
    try:
        import httpx
    except ImportError:
        return []

    month_code_map = {
        1: "jan", 2: "feb", 3: "mar", 4: "apr", 5: "may", 6: "jun",
        7: "jul", 8: "aug", 9: "sep", 10: "oct", 11: "nov", 12: "dec"
    }

    rows_out: list[dict] = []
    headers = {"User-Agent": _USER_AGENT}

    for dt in target_months:
        m_code = month_code_map.get(dt.month)
        if not m_code:
            continue
        url = f"https://portal.amfiindia.com/spages/am{m_code}{dt.year}repo.xls"
        try:
            resp = httpx.get(url, headers=headers, timeout=8)
            if resp.status_code != 200 or len(resp.content) < 5000:
                continue
            parsed = _parse_excel_content(resp.content, [dt])
            if parsed:
                rows_out.extend(parsed)
                logger.debug("Discovered portal link %s: got %d rows", url, len(parsed))
        except Exception as exc:
            logger.debug("Failed portal download %s: %s", url, exc)
            continue

    return rows_out

_AMFI_HOME = "https://www.amfiindia.com"
_STRAPI_PATHS = [
    "/api/amfi-monthly-data",
    "/api/monthly-data",
    "/api/category-wise-data",
]
_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
)


def _discover_strapi_host() -> Optional[str]:
    """
    Probe the AMFI homepage JS bundle to discover the Strapi CMS hostname.
    Returns the host URL (e.g. 'https://cms.amfiindia.com') or None.
    """
    try:
        import httpx
        resp = httpx.get(_AMFI_HOME, headers={"User-Agent": _USER_AGENT}, timeout=10, follow_redirects=True)
        # Look for CMS/API/backend subdomains in JS script tags
        pattern = re.compile(
            r'https?://(?:cms|api|backend|strapi|content)\.amfiindia\.com',
            re.IGNORECASE,
        )
        found = pattern.findall(resp.text)
        if found:
            host = found[0].rstrip("/")
            logger.debug("Discovered Strapi host: %s", host)
            return host
        # Also scan linked JS files for the pattern
        js_urls = re.findall(r'src="(/[^"]+\.js)"', resp.text)
        for js_path in js_urls[:5]:  # limit to first 5 JS files
            try:
                js_resp = httpx.get(
                    f"{_AMFI_HOME}{js_path}",
                    headers={"User-Agent": _USER_AGENT},
                    timeout=8,
                    follow_redirects=True,
                )
                found = pattern.findall(js_resp.text)
                if found:
                    host = found[0].rstrip("/")
                    logger.debug("Discovered Strapi host in JS bundle: %s", host)
                    return host
            except Exception:
                continue
    except Exception as exc:
        logger.debug("Strapi host discovery failed: %s", exc)
    return None


def _fetch_via_strapi(strapi_host: str, target_months: list[date]) -> list[dict]:
    """
    Attempt to pull category flow data from the Strapi REST API.
    Returns rows on success, empty list on failure.
    """
    try:
        import httpx
    except ImportError:
        return []

    rows = []
    for month_date in target_months:
        month_str = month_date.strftime("%Y-%m")
        for path in _STRAPI_PATHS:
            for eq_filter in (f"?filters[month_year][$eq]={month_str}&populate=*",
                              f"?filters[month][$eq]={month_str}&populate=*",
                              f"?month={month_str}"):
                url = f"{strapi_host}{path}{eq_filter}"
                try:
                    resp = httpx.get(
                        url, headers={"User-Agent": _USER_AGENT}, timeout=10, follow_redirects=True
                    )
                    if resp.status_code != 200:
                        continue
                    data = resp.json()
                    parsed = _parse_strapi_response(data, month_date)
                    if parsed:
                        rows.extend(parsed)
                        logger.debug("Strapi: got %d rows for %s", len(parsed), month_str)
                        break
                except Exception as exc:
                    logger.debug("Strapi attempt failed (%s): %s", url[:60], exc)
                    continue
            if any(r["report_month"] == month_date for r in rows):
                break
        time.sleep(0.3)  # polite rate limiting
    return rows


def _parse_strapi_response(data: dict, month_date: date) -> list[dict]:
    """Try to extract category rows from a Strapi JSON response."""
    # Strapi v4 wraps results in {"data": [{"id": N, "attributes": {...}}]}
    entries = []
    if isinstance(data, dict) and "data" in data:
        raw = data["data"]
        if isinstance(raw, list):
            entries = [e.get("attributes", e) for e in raw]
        elif isinstance(raw, dict):
            entries = [raw.get("attributes", raw)]
    elif isinstance(data, list):
        entries = data

    rows = []
    for entry in entries:
        # Try common field name patterns
        category = (
            entry.get("category_name") or entry.get("scheme_type") or
            entry.get("category") or entry.get("name") or ""
        )
        gross_purchase = _parse_amount(
            entry.get("gross_purchase") or entry.get("gross_purchase_cr") or
            entry.get("inflow") or entry.get("gross_inflow") or 0
        )
        gross_redemption = _parse_amount(
            entry.get("gross_redemption") or entry.get("gross_redemption_cr") or
            entry.get("outflow") or entry.get("gross_outflow") or 0
        )
        closing_aum = _parse_amount(
            entry.get("closing_aum") or entry.get("closing_aum_cr") or
            entry.get("aum") or 0
        )
        if not category or (gross_purchase == 0 and gross_redemption == 0 and closing_aum == 0):
            continue
        net_flow = gross_purchase - gross_redemption
        aum_safe = closing_aum if closing_aum != 0 else None
        rows.append({
            "report_month":        month_date,
            "category_name":       category.strip(),
            "subcategory_group":   _normalize_subcategory(category.strip()),
            "gross_purchase_cr":   gross_purchase,
            "gross_redemption_cr": gross_redemption,
            "net_flow_cr":         net_flow,
            "closing_aum_cr":      closing_aum,
            "flow_pct_of_aum":     (net_flow / aum_safe * 100) if aum_safe else 0.0,
        })
    return rows


# ── Layer 2: HTML scrape ──────────────────────────────────────────────────────

_AMFI_MONTHLY_URL = "https://www.amfiindia.com/research-information/amfi-monthly"
_AMFI_INDUSTRY_TRENDS = "https://www.amfiindia.com/research-information/industry-trends"


def _fetch_via_html_scrape(target_months: list[date]) -> list[dict]:
    """
    Scrape AMFI website HTML tables for category-wise flow data.
    Returns rows on success, empty list on failure.
    """
    try:
        import httpx
        from bs4 import BeautifulSoup
    except ImportError:
        logger.debug("httpx or beautifulsoup4 not installed; skipping HTML scrape layer")
        return []

    headers = {
        "User-Agent": _USER_AGENT,
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-US,en;q=0.9",
    }

    rows = []
    for url in (_AMFI_MONTHLY_URL, _AMFI_INDUSTRY_TRENDS):
        try:
            resp = httpx.get(url, headers=headers, timeout=15, follow_redirects=True)
            if resp.status_code != 200:
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            tables = soup.find_all("table")
            for table in tables:
                parsed = _parse_html_table(table, target_months)
                if parsed:
                    rows.extend(parsed)
        except Exception as exc:
            logger.debug("HTML scrape failed for %s: %s", url, exc)
            continue

    return rows


def _parse_html_table(table, target_months: list[date]) -> list[dict]:
    """
    Parse an HTML <table> element looking for AMFI category flow rows.
    Heuristic: look for rows with numeric ₹ Crore columns and a category label.
    """
    try:
        from bs4 import BeautifulSoup  # noqa: F401
    except ImportError:
        return []

    rows_out = []
    all_rows = table.find_all("tr")
    if len(all_rows) < 3:
        return []

    # Try to detect a month from the table header
    header_text = " ".join(th.get_text(strip=True) for th in (all_rows[0].find_all("th") or all_rows[0].find_all("td")))
    detected_month = _detect_month_from_header(header_text, target_months)

    for tr in all_rows[1:]:
        cells = [td.get_text(strip=True).replace("\xa0", " ") for td in tr.find_all("td")]
        if len(cells) < 4:
            continue
        # First cell should be a category label (text-heavy), remaining should be numeric
        category = cells[0].strip()
        if not category or category.isdigit() or len(category) < 5:
            continue
        numeric_cells = cells[1:]
        amounts = []
        for c in numeric_cells[:4]:
            try:
                amounts.append(_parse_amount(c))
            except Exception:
                amounts.append(0.0)
        if len(amounts) < 2:
            continue
        gross_purchase = amounts[0] if amounts[0] > 0 else 0.0
        gross_redemption = amounts[1] if len(amounts) > 1 and amounts[1] > 0 else 0.0
        closing_aum = amounts[3] if len(amounts) > 3 and amounts[3] > 0 else amounts[2] if len(amounts) > 2 else 0.0
        net_flow = gross_purchase - gross_redemption
        if gross_purchase == 0 and gross_redemption == 0 and closing_aum == 0:
            continue

        report_month = detected_month or (target_months[0] if target_months else date.today().replace(day=1))
        aum_safe = closing_aum if closing_aum != 0 else None
        rows_out.append({
            "report_month":        report_month,
            "category_name":       category,
            "subcategory_group":   _normalize_subcategory(category),
            "gross_purchase_cr":   gross_purchase,
            "gross_redemption_cr": gross_redemption,
            "net_flow_cr":         net_flow,
            "closing_aum_cr":      closing_aum,
            "flow_pct_of_aum":     (net_flow / aum_safe * 100) if aum_safe else 0.0,
        })
    return rows_out


def _detect_month_from_header(header_text: str, target_months: list[date]) -> Optional[date]:
    """Try to detect a YYYY-MM pattern from a table header string."""
    # e.g. "June 2026", "Jun-26", "2026-06"
    patterns = [
        (r"(\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})", "%B %Y"),
        (r"(\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2,4})", None),
        (r"(\d{4}-\d{2})", "%Y-%m"),
    ]
    for pattern, fmt in patterns:
        match = re.search(pattern, header_text, re.IGNORECASE)
        if match:
            try:
                if fmt:
                    dt = datetime.strptime(match.group(1), fmt)
                else:
                    raw = match.group(1)
                    # Handle Jan-26 / Jan-2026
                    dt = datetime.strptime(raw, "%b-%y") if len(raw.split("-")[1]) == 2 else datetime.strptime(raw, "%b-%Y")
                return dt.replace(day=1).date()
            except ValueError:
                continue
    return None


# ── Layer 3: Manual URL override (Excel/CSV download) ────────────────────────

def _fetch_via_manual_url(target_months: list[date]) -> list[dict]:
    """
    Download and parse an AMFI Excel/CSV file from AMFI_EXCEL_URL env var
    or from a local file in output/ (e.g. output/amfi_category_flows.xlsx).
    Handles both .xlsx (openpyxl) and .csv formats.
    """
    url_or_path = os.environ.get("AMFI_EXCEL_URL", "").strip() or os.environ.get("AMFI_FILE_PATH", "").strip()

    # Default fallback check for local files in output/ directory
    if not url_or_path:
        for candidate in ("output/amfi_category_flows.xlsx", "output/amfi_category_flows.csv",
                          "output/amfi_data.xlsx", "output/amfi_data.csv"):
            if os.path.exists(candidate):
                url_or_path = candidate
                logger.info("Found default local AMFI sheet: %s", candidate)
                break

    if not url_or_path:
        return []

    # If it's a local file path
    if os.path.exists(url_or_path):
        logger.info("Parsing local AMFI file: %s", url_or_path)
        try:
            with open(url_or_path, "rb") as f:
                content = f.read()
            if url_or_path.lower().endswith(".csv"):
                return _parse_csv_content(content, target_months)
            else:
                return _parse_excel_content(content, target_months)
        except Exception as exc:
            logger.warning("Failed to read local AMFI file %s: %s", url_or_path, exc)
            return []

    # Otherwise treat as HTTP URL
    logger.info("Fetching AMFI data from manual override URL: %s", url_or_path)
    try:
        import httpx
        resp = httpx.get(url_or_path, headers={"User-Agent": _USER_AGENT}, timeout=30, follow_redirects=True)
        if resp.status_code != 200:
            logger.warning("Manual URL returned HTTP %d", resp.status_code)
            return []
        content = resp.content
    except Exception as exc:
        logger.warning("Manual URL fetch failed: %s", exc)
        return []

    content_type = resp.headers.get("content-type", "").lower()
    if "csv" in content_type or url_or_path.lower().endswith(".csv"):
        return _parse_csv_content(content, target_months)
    else:
        return _parse_excel_content(content, target_months)


def _parse_excel_content(content: bytes, target_months: list[date]) -> list[dict]:
    """Parse AMFI Excel binary content using openpyxl."""
    try:
        import io
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed. Install with: pip install openpyxl")
        return []

    rows_out = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        # Try "Category-Wise Data" sheet first, then first sheet
        sheet_names = wb.sheetnames
        target_sheet = None
        for name in sheet_names:
            if "category" in name.lower() or "wise" in name.lower():
                target_sheet = wb[name]
                break
        if target_sheet is None and sheet_names:
            target_sheet = wb[sheet_names[0]]
        if target_sheet is None:
            return []

        # Detect header row and month
        detected_month = target_months[0] if target_months else date.today().replace(day=1)
        all_rows = list(target_sheet.iter_rows(values_only=True))

        for row in all_rows:
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not cells or not cells[0]:
                continue
            category = cells[0].strip()
            if not category or category.lower() in ("category", "scheme type", "total", "grand total", "none"):
                continue
            if category.replace(".", "").replace(",", "").isdigit():
                continue
            try:
                gross_purchase = _parse_amount(cells[1]) if len(cells) > 1 else 0.0
                gross_redemption = _parse_amount(cells[2]) if len(cells) > 2 else 0.0
                closing_aum = _parse_amount(cells[4]) if len(cells) > 4 else (
                    _parse_amount(cells[3]) if len(cells) > 3 else 0.0
                )
            except Exception:
                continue
            if gross_purchase == 0 and gross_redemption == 0 and closing_aum == 0:
                continue
            net_flow = gross_purchase - gross_redemption
            aum_safe = closing_aum if closing_aum != 0 else None
            rows_out.append({
                "report_month":        detected_month,
                "category_name":       category,
                "subcategory_group":   _normalize_subcategory(category),
                "gross_purchase_cr":   gross_purchase,
                "gross_redemption_cr": gross_redemption,
                "net_flow_cr":         net_flow,
                "closing_aum_cr":      closing_aum,
                "flow_pct_of_aum":     (net_flow / aum_safe * 100) if aum_safe else 0.0,
            })
    except Exception as exc:
        logger.warning("Excel parse failed: %s", exc)
    return rows_out


def _parse_csv_content(content: bytes, target_months: list[date]) -> list[dict]:
    """Parse AMFI CSV content."""
    import csv
    import io

    rows_out = []
    detected_month = target_months[0] if target_months else date.today().replace(day=1)
    try:
        reader = csv.reader(io.StringIO(content.decode("utf-8", errors="replace")))
        for row in reader:
            if not row or not row[0].strip():
                continue
            category = row[0].strip()
            if category.lower() in ("category", "scheme type", "total", "grand total"):
                continue
            try:
                gross_purchase = _parse_amount(row[1]) if len(row) > 1 else 0.0
                gross_redemption = _parse_amount(row[2]) if len(row) > 2 else 0.0
                closing_aum = _parse_amount(row[4]) if len(row) > 4 else (
                    _parse_amount(row[3]) if len(row) > 3 else 0.0
                )
            except Exception:
                continue
            if gross_purchase == 0 and gross_redemption == 0 and closing_aum == 0:
                continue
            net_flow = gross_purchase - gross_redemption
            aum_safe = closing_aum if closing_aum != 0 else None
            rows_out.append({
                "report_month":        detected_month,
                "category_name":       category,
                "subcategory_group":   _normalize_subcategory(category),
                "gross_purchase_cr":   gross_purchase,
                "gross_redemption_cr": gross_redemption,
                "net_flow_cr":         net_flow,
                "closing_aum_cr":      closing_aum,
                "flow_pct_of_aum":     (net_flow / aum_safe * 100) if aum_safe else 0.0,
            })
    except Exception as exc:
        logger.warning("CSV parse failed: %s", exc)
    return rows_out


# ── Public API ────────────────────────────────────────────────────────────────

def fetch_amfi_category_flows(months_back: int = 24) -> list[dict]:
    """
    Fetch AMFI Category-Wise Monthly Flows + AUM for the last `months_back` months.

    Tries four layers in order:
      1. Direct Discovered Portal SIF Excel links (portal.amfiindia.com)
      2. Strapi CMS API probe
      3. HTML scrape (httpx + BeautifulSoup)
      4. Manual URL override / local sheet file (AMFI_EXCEL_URL / AMFI_FILE_PATH)

    Returns a list of dicts with keys:
        report_month, category_name, subcategory_group,
        gross_purchase_cr, gross_redemption_cr, net_flow_cr,
        closing_aum_cr, flow_pct_of_aum

    Returns [] if all layers fail — never raises.
    """
    target_months = _month_dates(months_back)
    logger.info(
        "Fetching AMFI category flows for %d months back (%s → %s)",
        months_back,
        target_months[-1].strftime("%b %Y"),
        target_months[0].strftime("%b %Y"),
    )

    # ── Layer 1: Direct Discovered Portal SIF Excel links ─────────────────
    logger.info("Probing direct AMFI portal statistical files (2019-2026)…")
    rows = _fetch_via_portal_statistical_reports(target_months)
    if rows:
        logger.info("Direct Portal layer returned %d rows across %d months", len(rows),
                    len({r["report_month"] for r in rows}))
        return _deduplicate(rows)
    logger.debug("Direct Portal layer returned no rows")

    # ── Layer 2: HTML scrape ───────────────────────────────────────────────
    logger.info("Trying HTML scrape fallback (amfiindia.com tables)…")
    rows = _fetch_via_html_scrape(target_months)
    if rows:
        logger.info("HTML scrape returned %d rows across %d months", len(rows),
                    len({r["report_month"] for r in rows}))
        return _deduplicate(rows)
    logger.debug("HTML scrape returned no rows")

    # ── Layer 3: Manual URL override ───────────────────────────────────────
    rows = _fetch_via_manual_url(target_months)
    if rows:
        logger.info("Manual URL layer returned %d rows", len(rows))
        return _deduplicate(rows)

    # ── All layers failed ──────────────────────────────────────────────────
    logger.warning(
        "[WARNING] AMFI category flows unavailable — all 3 fetch layers failed. "
        "Set AMFI_EXCEL_URL in .env to a direct download link as a manual fallback. "
        "The import pipeline will continue without this data."
    )
    return []


def _deduplicate(rows: list[dict]) -> list[dict]:
    """Remove duplicate (report_month, category_name) pairs, keeping last seen."""
    seen: dict[tuple, dict] = {}
    for row in rows:
        key = (row["report_month"], row["category_name"])
        seen[key] = row
    result = sorted(seen.values(), key=lambda r: (r["report_month"], r["category_name"]))
    logger.debug("After dedup: %d unique (month, category) rows", len(result))
    return result
